from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = Workbook()
wb = load_workbook('data.xlsx')
ws = wb.active
def create_database_nd(pe,fe,rne,a1e,a2e,b1e,b2e,tc):

    ws["A1"] = "Turbine / Compressor"
    ws["B1"] = "ψ"
    ws["C1"] = "φ"
    ws["D1"] = "Rn"
    ws["E1"] = "α1"
    ws["F1"] = "α2"
    ws["G1"] = "β1"
    ws["H1"] = "β2"

    j = ws.max_row + 1
    if tc == 0:
        name = 'Turbine'
    else:
        name = 'Compressor'

    ws["A" + str(j)] = name
    ws["B" + str(j)] = str(pe)
    ws["C" + str(j)] = str(fe)
    ws["D" + str(j)] = str(rne)
    ws["E" + str(j)] = str(a1e)
    ws["F" + str(j)] = str(a2e)
    ws["G" + str(j)] = str(b1e)
    ws["H" + str(j)] = str(b2e)

    wb.save('data.xlsx')
# def create_database_d():
#     ws["I1"] = ""
#     ws["B1"] = "φ"
#     ws["C1"] = "Rn"
#     ws["D1"] = "α1"
#     ws["E1"] = "α2"
#     ws["F1"] = "β1"
#     ws["G1"] = "β2"
#
#     j = ws.max_row + 1
#     ws["A" + str(j)] = str(pe)
#     ws["B" + str(j)] = str(fe)
#     ws["C" + str(j)] = str(rne)
#     ws["D" + str(j)] = str(a1e)
#     ws["E" + str(j)] = str(a2e)
#     ws["F" + str(j)] = str(b1e)
#     ws["G" + str(j)] = str(b2e)
# save new file
    wb.save('data.xlsx')